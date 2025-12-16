import numpy as np
import pandas as pd
import os
import scipy.io
from openpyxl import Workbook

def generate_multi_group_data():
    print("=== 多组随机数据生成器 ===")
    try:
        num_groups_str = input("请输入要生成的组数 (Number of groups): ")
        if not num_groups_str:
            print("输入不能为空喵！")
            return
        num_groups = int(num_groups_str)
        
        all_data = {}
        max_length = 0
        
        group_configs = []

        for i in range(num_groups):
            print(f"\n--- 设置第 {i+1} 组数据 ---")
            name = input(f"请输入第 {i+1} 组的名称 (默认为 Group_{i+1}): ")
            if not name:
                name = f"Group_{i+1}"
            
            mean_str = input(f"请输入 {name} 的期望平均数 (Mean): ")
            mean = float(mean_str)
            
            std_str = input(f"请输入 {name} 的期望标准差 (Std Dev): ")
            std = float(std_str)
            
            count_str = input(f"请输入 {name} 的数据量 (Count): ")
            count = int(count_str)
            
            # 生成数据
            data = np.random.normal(loc=mean, scale=std, size=count)
            data = np.round(data, 1) # 保留一位小数
            
            all_data[name] = data
            if count > max_length:
                max_length = count
            
            group_configs.append({
                'name': name,
                'mean': mean,
                'std': std,
                'count': count
            })

        # 保存到 Excel (自定义格式: 标题 -> 10个数据一行 -> 空行)
        excel_filename = 'multi_group_data.xlsx'
        current_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(current_dir, excel_filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Random Data"
        
        current_row = 1
        for name, data in all_data.items():
            # 写入标题
            ws.cell(row=current_row, column=1, value=name)
            current_row += 1
            
            # 写入数据 (每行10个)
            col_idx = 1
            for value in data:
                ws.cell(row=current_row, column=col_idx, value=value)
                col_idx += 1
                if col_idx > 10:
                    col_idx = 1
                    current_row += 1
            
            # 如果最后一行没填满，换行
            if col_idx != 1:
                current_row += 1
            
            # 空一行
            current_row += 1
            
        wb.save(excel_path)
        print(f"\n成功保存 Excel 文件至: {excel_path}")
        
        # 保存到 .mat (统一数据格式)
        mat_filename = 'multi_group_data.mat'
        mat_path = os.path.join(current_dir, mat_filename)
        
        # 准备 mat 数据
        # 注意: MATLAB 变量名不能包含空格，这里简单替换一下
        mat_data = {}
        for name, data in all_data.items():
            safe_name = name.replace(' ', '_')
            # MATLAB 变量名必须以字母开头
            if not safe_name or not safe_name[0].isalpha():
                safe_name = "Group_" + safe_name
            mat_data[safe_name] = data
        
        scipy.io.savemat(mat_path, mat_data)
        print(f"成功保存 MAT 文件至: {mat_path}")
        
        print("\n数据生成完毕！")
        print("生成的组信息:")
        for config in group_configs:
            print(f"  - {config['name']}: Mean={config['mean']}, Std={config['std']}, Count={config['count']}")

    except ValueError as e:
        print(f"输入错误: {e}")
    except Exception as e:
        print(f"发生未知错误: {e}")

if __name__ == "__main__":
    generate_multi_group_data()
