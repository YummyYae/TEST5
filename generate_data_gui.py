import sys
import os
import numpy as np
from openpyxl import Workbook
from PyQt6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QPushButton, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QMessageBox, QGroupBox, QFrame)
from PyQt6.QtCore import Qt

class RandomDataGeneratorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.groups = []
        self.initUI()
        self.init_default_groups()

    def initUI(self):
        self.setWindowTitle('随机数据生成器 (PyQt6版)')
        self.setGeometry(100, 100, 700, 600)

        # 主布局
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        # === 输入区域 ===
        input_group = QGroupBox("添加新数据组")
        input_layout = QVBoxLayout()
        input_group.setLayout(input_layout)
        
        # 表单布局 (使用 HBoxLayouts 模拟网格，或者直接用 QGridLayout)
        # 第一行
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("组名称:"))
        self.name_input = QLineEdit()
        row1.addWidget(self.name_input)
        
        row1.addWidget(QLabel("期望平均数 (Mean):"))
        self.mean_input = QLineEdit("0.0")
        row1.addWidget(self.mean_input)
        input_layout.addLayout(row1)

        # 第二行
        row2 = QHBoxLayout()
        row2.addWidget(QLabel("标准差 (Std Dev):"))
        self.std_input = QLineEdit("1.0")
        row2.addWidget(self.std_input)
        
        row2.addWidget(QLabel("数据量 (Count):"))
        self.count_input = QLineEdit("50")
        row2.addWidget(self.count_input)
        input_layout.addLayout(row2)

        # 添加按钮
        self.add_btn = QPushButton("添加组")
        self.add_btn.clicked.connect(self.add_group)
        input_layout.addWidget(self.add_btn)

        main_layout.addWidget(input_group)

        # === 列表区域 ===
        list_group = QGroupBox("已添加的数据组")
        list_layout = QVBoxLayout()
        list_group.setLayout(list_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["名称", "平均数", "标准差", "数量"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        # 允许编辑，以便微调
        # self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers) 
        self.table.itemChanged.connect(self.on_item_changed)
        list_layout.addWidget(self.table)

        # 删除按钮
        self.del_btn = QPushButton("删除选中组")
        self.del_btn.clicked.connect(self.delete_group)
        list_layout.addWidget(self.del_btn, alignment=Qt.AlignmentFlag.AlignRight)

        main_layout.addWidget(list_group)

        # === 生成区域 ===
        action_frame = QFrame()
        action_layout = QVBoxLayout()
        action_frame.setLayout(action_layout)
        
        self.generate_btn = QPushButton("生成 Excel")
        self.generate_btn.setFixedHeight(40)
        self.generate_btn.clicked.connect(self.generate_files)
        action_layout.addWidget(self.generate_btn)
        
        self.status_label = QLabel("准备就绪")
        self.status_label.setFrameStyle(QFrame.Shape.Panel | QFrame.Shadow.Sunken)
        action_layout.addWidget(self.status_label)

        main_layout.addWidget(action_frame)

    def add_group(self):
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "错误", "请输入组名称")
            return
        
        try:
            mean = float(self.mean_input.text())
            std = float(self.std_input.text())
            count = int(self.count_input.text())
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的数字")
            return

        if count <= 0:
            QMessageBox.warning(self, "错误", "数据量必须大于0")
            return

        # 检查重名
        for g in self.groups:
            if g['name'] == name:
                QMessageBox.warning(self, "错误", f"组名 '{name}' 已存在")
                return

        group_data = {
            'name': name,
            'mean': mean,
            'std': std,
            'count': count
        }
        self.groups.append(group_data)
        
        # 更新表格
        row_idx = self.table.rowCount()
        self.table.insertRow(row_idx)
        self.table.setItem(row_idx, 0, QTableWidgetItem(name))
        self.table.setItem(row_idx, 1, QTableWidgetItem(str(mean)))
        self.table.setItem(row_idx, 2, QTableWidgetItem(str(std)))
        self.table.setItem(row_idx, 3, QTableWidgetItem(str(count)))
        
        # 清空输入
        self.name_input.clear()
        self.status_label.setText(f"已添加组: {name}")

    def delete_group(self):
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        # 从后往前删，避免索引错乱
        for index in sorted(selected_rows, reverse=True):
            row = index.row()
            name = self.table.item(row, 0).text()
            
            # 从数据列表中移除
            self.groups = [g for g in self.groups if g['name'] != name]
            
            # 从表格中移除
            self.table.removeRow(row)
            self.status_label.setText(f"已删除组: {name}")

    def init_default_groups(self):
        default_names = [
            "教二北逆时针", "教二南逆时针", "教二二层走廊", "教二二层教室",
            "教二四层走廊", "教二四层教室", "教三北逆时针", "教三南逆时针",
            "教三二层走廊", "教三二层教室", "教三四层走廊", "教三四层教室"
        ]
        
        self.table.blockSignals(True)
        for name in default_names:
            mean = 45.0
            std = 3.0
            count = 50
            
            group_data = {
                'name': name,
                'mean': mean,
                'std': std,
                'count': count
            }
            self.groups.append(group_data)
            
            row_idx = self.table.rowCount()
            self.table.insertRow(row_idx)
            self.table.setItem(row_idx, 0, QTableWidgetItem(name))
            self.table.setItem(row_idx, 1, QTableWidgetItem(str(mean)))
            self.table.setItem(row_idx, 2, QTableWidgetItem(str(std)))
            self.table.setItem(row_idx, 3, QTableWidgetItem(str(count)))
        self.table.blockSignals(False)

    def on_item_changed(self, item):
        row = item.row()
        col = item.column()
        
        if row < 0 or row >= len(self.groups):
            return
            
        try:
            text = item.text()
            if col == 0: # Name
                self.groups[row]['name'] = text
            elif col == 1: # Mean
                self.groups[row]['mean'] = float(text)
            elif col == 2: # Std
                self.groups[row]['std'] = float(text)
            elif col == 3: # Count
                self.groups[row]['count'] = int(text)
        except ValueError:
            pass

    def generate_files(self):
        if not self.groups:
            QMessageBox.warning(self, "警告", "请先添加至少一组数据")
            return

        try:
            all_data = {}
            
            # 生成数据
            for group in self.groups:
                raw_data = np.random.normal(loc=group['mean'], scale=group['std'], size=group['count'])
                
                # === 数据映射处理 ===
                # 复制一份数据进行处理，避免原地修改造成的逻辑混淆
                processed_data = raw_data.copy()
                
                # 1. 小于 0 -> 映射到 20
                processed_data[raw_data < 0] = 20
                
                # 2. 0 <= data < 20 -> 对数映射到 20-25
                # 公式: y = 20 + 5 * ln(x + 1) / ln(21)
                mask_0_20 = (raw_data >= 0) & (raw_data < 20)
                if np.any(mask_0_20):
                    processed_data[mask_0_20] = 20 + 5 * np.log(raw_data[mask_0_20] + 1) / np.log(21)
                
                # 3. 20 <= data < 30 -> 线性映射到 25-30
                # 公式: y = 25 + (x - 20) * 0.5
                mask_20_30 = (raw_data >= 20) & (raw_data < 30)
                if np.any(mask_20_30):
                    processed_data[mask_20_30] = 25 + (raw_data[mask_20_30] - 20) * 0.5
                
                # --- 上界处理 (对称逻辑，确保数据落在 20-60 之间) ---
                # 假设核心区间为 [30, 50] 保持不变
                
                # 4. 50 <= data < 60 -> 线性映射到 50-55
                # 公式: y = 50 + (x - 50) * 0.5
                mask_50_60 = (raw_data >= 50) & (raw_data < 60)
                if np.any(mask_50_60):
                    processed_data[mask_50_60] = 50 + (raw_data[mask_50_60] - 50) * 0.5
                    
                # 5. 60 <= data < 80 -> 对数映射到 55-60
                # 公式: y = 55 + 5 * ln(x - 60 + 1) / ln(21)
                mask_60_80 = (raw_data >= 60) & (raw_data < 80)
                if np.any(mask_60_80):
                    processed_data[mask_60_80] = 55 + 5 * np.log(raw_data[mask_60_80] - 60 + 1) / np.log(21)
                    
                # 6. 大于等于 80 -> 映射到 60
                processed_data[raw_data >= 80] = 60
                
                data = np.round(processed_data, 1) # 保留一位小数
                all_data[group['name']] = data

            # === 保存到 Excel ===
            excel_filename = 'data.xlsx'
            # 获取当前脚本所在目录
            # 注意：如果是打包后的exe，路径获取方式可能不同，这里假设是脚本运行
            current_dir = os.path.dirname(os.path.abspath(__file__))
            excel_path = os.path.join(current_dir, excel_filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Random Data"
            
            # 按列写入数据
            col_idx = 1
            for name, data in all_data.items():
                # 写入标题 (第一行)
                ws.cell(row=1, column=col_idx, value=name)
                
                # 写入数据 (从第二行开始)
                for row_idx, value in enumerate(data, start=2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                col_idx += 1
                
            wb.save(excel_path)
            
            QMessageBox.information(self, "成功", f"文件已生成！\nExcel: {excel_filename}")
            self.status_label.setText("生成完毕")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成过程中发生错误:\n{str(e)}")
            self.status_label.setText("生成失败")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RandomDataGeneratorApp()
    ex.show()
    sys.exit(app.exec())
